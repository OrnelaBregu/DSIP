import pdfplumber
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook


def extract_main_text(pdf_path):
    """Extract text from the main content using pdfplumber."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_widget_text(pdf_path):
    """Extract text from form fields/widgets using PyMuPDF."""
    doc = fitz.open(pdf_path)
    widget_text = ""
    for page in doc:
        widgets = page.widgets()
        if widgets:
            for widget in widgets:
                value = widget.field_value
                if value:
                    widget_text += f"{widget.field_name or ''}: {value}\n"
    return widget_text


def extract_fields(text):
    """
    Extract fields from given text using simple line-by-line parsing.
    Fields are expected to be in the format 'Field Name: Value'.
    """
    fields = {}
    for line in text.splitlines():
        line = line.strip()
        if ":" in line:
            key, value = line.split(":", 1)
            fields[key.strip()] = value.strip()
    return fields


def parse_pdf_data(pdf_path):
    """
    Extract text from both the main content and widget fields,
    then combine the extracted fields (with widget values taking precedence)
    to build the final data dictionary matching our Excel template.
    """
    main_text = extract_main_text(pdf_path)
    widget_text = extract_widget_text(pdf_path)
    main_fields = extract_fields(main_text)
    widget_fields = extract_fields(widget_text)

    # Default values based on your Excel template
    default_data = {
        "Student Name": "Ornela",
        "Thesis Title": "LLM",
        "Student ID": "26898580",
        "Department": "CIISE",
        "TH05": "TH05",
        "Thesis Defence Date": "MArch 20, 2025",
        "Oral Defence": "Yes",
        "Room": "309",
        "Thesis Ranking": "Outstanding",
        "Coded on SIS": "",
        "Embargo Date": "",
        "Examining Committee decision": "Accepted"
    }

    mapping = {
        "Student Name": "Student Name",
        "Thesis Title": "Thesis Title",
        "Student ID": "Student ID",
        "Department": "Department",
        "Defence date": "Thesis Defence Date",
        "Room": "Room",
        "Thesis ranking": "Thesis Ranking",
        "Decision": "Examining Committee decision",
        "Oral Defence": "Oral Defence",
        "Date": "Coded on SIS"
    }

    combined_data = default_data.copy()

    def update_fields(source_fields):
        for key, value in source_fields.items():
            for map_key, target_field in mapping.items():
                if key.lower() == map_key.lower() and value:
                    combined_data[target_field] = value

    update_fields(main_fields)
    update_fields(widget_fields)

    return combined_data


def append_data_to_excel(data, output_excel, template_path):
    """
    Append a new row of extracted data into the macro-enabled workbook.
    If the output file doesn't exist, it will be created using the template.
    """
    # Define the desired column order.
    columns = [
        "Student Name", "Thesis Title", "Student ID", "Department",
        "TH05", "Thesis Defence Date", "Oral Defence", "Room",
        "Coded on SIS", "Embargo Date", "Examining Committee decision", "Thesis Ranking"
    ]

    # Try to load the existing workbook; if not, load the template.
    try:
        wb = load_workbook(output_excel, keep_vba=True)
        ws = wb.active
    except FileNotFoundError:
        wb = load_workbook(template_path, keep_vba=True)
        ws = wb.active

    # Create a list representing a new row based on the desired column order.
    new_row = [data.get(col, "") for col in columns]

    # Append the row to the active worksheet.
    ws.append(new_row)

    wb.save(output_excel)
    print("Data appended to", output_excel)


if __name__ == "__main__":
    # Process one PDF at a time.
    # pdf_path = r"C:\Users\umroot\Downloads\DSIP\MA Committee Report - Youssef Maghrebi_40259660.pdf"
    # pdf_path = r"C:\Users\umroot\Downloads\DSIP\MA Committee Report -Ornela_Bregu_26898580.pdf"
    # output_excel = r"C:\Users\umroot\Downloads\DSIP\ExtractPdfData.xlsm"
    # template_path = r"C:\Users\umroot\Downloads\DSIP\Template.xlsm"
    pdf_path = r"C:\Users\umroot\PycharmProjects\DSIP\phd-committee-formOmar40187210.pdf"

    output_excel = r"C:\Users\umroot\PycharmProjects\DSIP\ExtractPdfData.xlsm"
    template_path = r"C:\Users\umroot\PycharmProjects\DSIP\Template.xlsm"

    data = parse_pdf_data(pdf_path)
    append_data_to_excel(data, output_excel, template_path)
