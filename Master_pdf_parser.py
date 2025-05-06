import logging
from pathlib import Path
import pdfplumber
import fitz  # PyMuPDF
from PyPDF2 import PdfReader
from dateutil import parser as date_parser
from openpyxl import load_workbook
import json

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(Y%m%d %H:%M:%S) [%(levelname)s] %(message)s"
)

# --- Configuration ------------------------------------
COLUMNS = [
    "Student Name", "Thesis Title", "Student ID", "Department",
    "TH05", "Thesis Defence Date", "Oral Defence", "Room",
    "Coded on SIS", "Embargo Date", "Examining Committee decision", "Thesis Ranking"
]

FIELD_MAPPING = {
    "student name": "Student Name",
    "thesis title": "Thesis Title",
    "student id": "Student ID",
    "department": "Department",
    "defence date": "Thesis Defence Date",
    "oral defence": "Oral Defence",
    "room": "Room",
    "coded on sis": "Coded on SIS",
    "embargo date": "Embargo Date",
    "decision": "Examining Committee decision",
    "thesis ranking": "Thesis Ranking",
}

REQUIRED_FIELDS = ["Student Name", "Student ID"]

# Map decision choice5 to ACCEPTED
DECISION_MAP = {"/choice5": "ACCEPTED"}

# Map ranking choices
RANKING_MAP = {
    "/choice7": "Outstanding",
    "/choice8": "Excellent",
    "/choice9": "Very Good",
    "/choice10": "Satisfactory",
    "/choice11": "Non satisfactory"
}

# --- Extraction Helpers -------------------------------

def extract_main_text(pdf_path: Path) -> str:
    text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if page_text := page.extract_text():
                text.append(page_text)
    return "\n".join(text)


def extract_widget_fields(pdf_path: Path) -> dict:
    fields = {}
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            for widget in page.widgets() or []:
                if val := widget.field_value:
                    fields[widget.field_name or widget.field_alias] = val
    except Exception:
        logging.debug("PyMuPDF parse failed for %s", pdf_path.name)
    try:
        reader = PdfReader(str(pdf_path))
        for name, info in (reader.get_fields() or {}).items():
            if info.value:
                fields[name] = info.value
    except Exception:
        logging.debug("PyPDF2 parse failed for %s", pdf_path.name)
    return fields


def extract_key_values(text: str) -> dict:
    kv = {}
    for line in text.splitlines():
        if ":" in line:
            key, val = map(str.strip, line.split(":", 1))
            if key and val:
                kv[key.lower()] = val
    return kv


def normalize_date(value: str) -> str:
    try:
        dt = date_parser.parse(value)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return value

# --- Main parsing logic -------------------------------

def parse_pdf(pdf_path: Path) -> dict:
    logging.info("Parsing %s", pdf_path.name)
    main_text = extract_main_text(pdf_path)
    widget_fields = extract_widget_fields(pdf_path)
    main_kv = extract_key_values(main_text)

    combined = {**main_kv, **{k.lower(): v for k, v in widget_fields.items()}}
    record = {col: "" for col in COLUMNS}

    for raw_key, raw_val in combined.items():
        if col := FIELD_MAPPING.get(raw_key):
            val = normalize_date(raw_val) if "date" in col.lower() else raw_val
            record[col] = val

    # Map decision codes
    dec = record.get("Examining Committee decision", "").lower()
    if dec in DECISION_MAP:
        record["Examining Committee decision"] = DECISION_MAP[dec]

    # Map ranking codes; leave empty if no match
    raw_rank = record.get("Thesis Ranking", "").lower()
    record["Thesis Ranking"] = RANKING_MAP.get(raw_rank, "")

    return record

# --- Excel output -------------------------------------

def write_records_to_excel(records: list[dict], output_path: Path, template_path: Path):
    if output_path.exists():
        wb = load_workbook(output_path, keep_vba=True)
    else:
        wb = load_workbook(template_path, keep_vba=True)
    ws = wb.active
    for rec in records:
        ws.append([rec[col] for col in COLUMNS])
    wb.save(output_path)
    logging.info("Wrote %d records to %s", len(records), output_path.name)

# --- Orchestration ------------------------------------

def main():
    pdf_folder = Path("/Users/ornelabregu/PycharmProjects/DSIP/FORMS")
    output_excel = Path("/Users/ornelabregu/PycharmProjects/DSIP/ExtractPdfData.xlsm")
    template_file = Path("/Users/ornelabregu/PycharmProjects/DSIP/Template.xlsm")
    error_log = pdf_folder / "errors.json"

    valid_records = []
    failed_entries = []

    for pdf in pdf_folder.glob("*.pdf"):
        try:
            rec = parse_pdf(pdf)
            missing = [f for f in REQUIRED_FIELDS if not rec.get(f)]
            if missing:
                failed_entries.append({"file": pdf.name, "errors": [f"Missing required fields: {', '.join(missing)}"]})
            else:
                valid_records.append(rec)
        except Exception as e:
            failed_entries.append({"file": pdf.name, "errors": [str(e)]})

    if valid_records:
        write_records_to_excel(valid_records, output_excel, template_file)
    else:
        logging.warning("No valid records to write")

    if failed_entries:
        with open(error_log, 'w') as f:
            json.dump(failed_entries, f, indent=2)
        logging.info("Wrote error log to %s", error_log.name)

if __name__ == "__main__":
    main()
